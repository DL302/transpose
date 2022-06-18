from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
import sys

if len(sys.argv) == 2:
    filePath = sys.argv[1]
else:
    filePath = input("Enter Excel Workbook path: ")

wb = load_workbook(filePath)
print(f"Available sheets: {wb.sheetnames}")
ws = wb[input("Enter sheet name [default \"transpose\"]: ") or "transpose"]

geneNameCol = column_index_from_string(
    input("Enter category name column letter [default B]: ") or "B")
dataCol = column_index_from_string(
    input("Enter data column letter [default I]: ") or "I")
rowStart = int(input("Enter starting row number [default 2]: ") or 2)
rowEnd = int(input("Enter ending row number: "))

# this is how many rows to transpose at once
interval = int(input("Enter blocksize: "))

destRow = int(input("Enter destination starting row: "))
destCol = column_index_from_string(input("Enter destination starting col: "))

input("Close your workbook, then press Enter here: ")


def transpose(minRow, maxRow, minCol, maxCol, outRow, outCol):
    for row in range(minRow, maxRow + 1):
        for col in range(minCol, maxCol + 1):
            ws.cell(outRow + col - minCol, outCol + row -
                    minRow).value = ws.cell(row, col).value


for currentRow in range(rowStart, rowEnd + 1, interval):
    # get category name & copy paste it to transposed data
    ws.cell(destRow, destCol).value = ws.cell(currentRow, geneNameCol).value
    # transpose the data assocated with that category name
    transpose(currentRow, currentRow + interval - 1,
              dataCol, dataCol, destRow, destCol + 1)
    # for each transposition, the next one will occurr on the next row
    destRow += 1

wb.save(filePath)
print("done")
